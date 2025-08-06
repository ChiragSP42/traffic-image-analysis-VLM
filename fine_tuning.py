import boto3
import os
#%%
import pandas as pd

df = pd.read_csv("response_with_filenames.csv")

input_df = df.loc[:, ["File_name"]].copy()
input_df["File_name"] = input_df["File_name"]+".jpg"

input_df["Description"] = "Color: "+df["Response_colors"] \
                        + ", Make: " + df["Response_makes"] \
                        + ", Model: " + df["Response_models"] \
                        + ", Year: " + df["Response_years"] \
                        + ", Car Type: " + df["Response_car_type"] \
                        + ", Unique Identifiers: " + df["Response_identifiers"].apply(lambda x: str(x).replace("[", "").replace("]", "").replace("'", ""))

input_df.to_csv("input.csv", index=False)

# %%
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Load your CSV file
csv_file = 'input.csv'
df = pd.read_csv(csv_file)

# Create an Excel workbook and select the active worksheet
wb = Workbook()
ws = wb.active

# Write DataFrame to Excel sheet
for r_idx, row in enumerate(df.itertuples(index=False, name=None), 1):
    for c_idx, value in enumerate(row, 1):
        ws.cell(row=r_idx, column=c_idx, value=value)

# List of image file paths, indexed by row (must match order & length of your dataframe)
image_paths = ['path_to_image1.png', 'path_to_image2.png', ...]  # one per row

# The column index where images will be inserted (after last data column)
img_col = df.shape[1] + 1

# Insert each image into its respective row
for i, img_path in enumerate(image_paths, 1):
    img = Image(img_path)
    img.anchor = ws.cell(row=i, column=img_col).coordinate  # Attach image to cell
    ws.add_image(img)

# Save the workbook
wb.save('output_with_images.xlsx')

